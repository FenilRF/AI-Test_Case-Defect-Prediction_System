"""
Export Utilities
-----------------
Converts test-case records into downloadable CSV, JSON, and Excel payloads.
"""

import csv
import io
import json
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def test_cases_to_csv(test_cases: List[Dict[str, Any]]) -> str:
    """
    Convert a list of test-case dicts to a CSV string.

    Parameters
    ----------
    test_cases : list[dict]
        Each dict must have keys: test_id, module_name, scenario,
        test_type, test_level, expected_result, priority.

    Returns
    -------
    str
        CSV-formatted string ready for download.
    """
    if not test_cases:
        return ""

    fieldnames = ["test_id", "module_name", "scenario", "test_type", "test_level", "expected_result", "priority"]
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for tc in test_cases:
        writer.writerow(tc)
    return buffer.getvalue()


def test_cases_to_json(test_cases: List[Dict[str, Any]]) -> str:
    """
    Convert a list of test-case dicts to a pretty-printed JSON string.
    """
    return json.dumps(test_cases, indent=2, default=str)


# ── Excel Export ─────────────────────────────────────────────

# Color scheme for test types
_TYPE_FILLS = {
    "positive":   PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "negative":   PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
    "boundary":   PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "edge":       PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    "security":   PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
}

_LEVEL_FILLS = {
    "Unit":        PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "Integration": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    "System":      PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "UAT":         PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
}

_HEADER_FILL = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
_BODY_FONT   = Font(size=10, name="Calibri")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def test_cases_to_excel(test_cases: List[Dict[str, Any]]) -> bytes:
    """
    Convert a list of test-case dicts to a styled Excel workbook (bytes).

    Returns bytes suitable for StreamingResponse / file download.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # ── Headers ─────────────────────────────────────────
    headers = ["#", "Module", "Scenario", "Test Type", "Test Level", "Expected Result", "Priority"]
    col_widths = [6, 18, 50, 14, 14, 50, 10]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze the header row
    ws.freeze_panes = "A2"

    # ── Data Rows ───────────────────────────────────────
    for row_idx, tc in enumerate(test_cases, start=2):
        row_data = [
            f"TC_{row_idx - 1:03d}",
            tc.get("module_name", ""),
            tc.get("scenario", ""),
            tc.get("test_type", ""),
            tc.get("test_level", "Unit"),
            tc.get("expected_result", ""),
            tc.get("priority", "P3"),
        ]

        test_type = tc.get("test_type", "").lower()
        test_level = tc.get("test_level", "Unit")
        type_fill = _TYPE_FILLS.get(test_type)
        level_fill = _LEVEL_FILLS.get(test_level)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [3, 6]))

            # Apply type-based fill to the "Test Type" column
            if col_idx == 4 and type_fill:
                cell.fill = type_fill
            # Apply level-based fill to the "Test Level" column
            if col_idx == 5 and level_fill:
                cell.fill = level_fill

    # ── Summary Sheet ───────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 15

    summary_title = ws_summary.cell(row=1, column=1, value="Test Case Summary")
    summary_title.font = Font(bold=True, size=14, name="Calibri")
    ws_summary.merge_cells("A1:B1")

    # Count by type
    type_counts = {}
    level_counts = {}
    priority_counts = {}
    for tc in test_cases:
        t = tc.get("test_type", "unknown")
        l = tc.get("test_level", "Unit")
        p = tc.get("priority", "P3")
        type_counts[t] = type_counts.get(t, 0) + 1
        level_counts[l] = level_counts.get(l, 0) + 1
        priority_counts[p] = priority_counts.get(p, 0) + 1

    row = 3
    ws_summary.cell(row=row, column=1, value="By Test Type").font = Font(bold=True, size=11)
    row += 1
    for t, c in sorted(type_counts.items()):
        ws_summary.cell(row=row, column=1, value=t.capitalize())
        ws_summary.cell(row=row, column=2, value=c)
        row += 1

    row += 1
    ws_summary.cell(row=row, column=1, value="By Test Level").font = Font(bold=True, size=11)
    row += 1
    for l, c in sorted(level_counts.items()):
        ws_summary.cell(row=row, column=1, value=l)
        ws_summary.cell(row=row, column=2, value=c)
        row += 1

    row += 1
    ws_summary.cell(row=row, column=1, value="By Priority").font = Font(bold=True, size=11)
    row += 1
    for p, c in sorted(priority_counts.items()):
        ws_summary.cell(row=row, column=1, value=p)
        ws_summary.cell(row=row, column=2, value=c)
        row += 1

    ws_summary.cell(row=row + 1, column=1, value=f"Total: {len(test_cases)}").font = Font(bold=True)

    # ── Serialize to bytes ──────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

